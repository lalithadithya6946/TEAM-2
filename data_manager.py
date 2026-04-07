import sqlite3, os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DataManager:
    def __init__(self, db_path="students_data/students.db"):
        self.db_path = db_path
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._init_db()

    def validate_reg_no(self, reg_no):
        """Validate registration number format for university students
        
        Supports multiple formats:
        - 11 digits: 99YYNNNXXXX (e.g., 99220041142)
        - 10 digits: 99YYNNNXXX (e.g., 9922004114) 
        - 9 digits: 99YYNNNXX (e.g., 992200411)
        
        Where:
        - 99 = Fixed prefix
        - YY = enrollment year (23-26 for current valid range)
        - NNN = Department code
        - XXXX/XXX/XX = Student number (1-9999/1-999/1-99 respectively)
        """
        if not reg_no:
            return False, "Registration number is required"
        
        # Allow faculty accounts (e.g., PSBCSE)
        if reg_no.upper() == "PSBCSE":
            return True, ""
            
        # Check if it's numeric and starts with 99
        if not reg_no.isdigit() or not reg_no.startswith('99'):
            return False, "Registration number must be numeric and start with 99"
        
        length = len(reg_no)
        if length not in [9, 10, 11]:
            return False, "Registration number must be 9, 10, or 11 digits"
        
        # Validate ranges based on length
        try:
            year = int(reg_no[2:4])  # YY from position 2-3
            
            current_year = datetime.now().year % 100  # Get last 2 digits of current year
            if year > current_year or year < (current_year - 5):  # Allow up to 5 years old
                return False, f"Invalid year in registration number. Must be between {current_year-5:02d} and {current_year:02d}"
            
            if length == 11:
                # Format: 99YYNNNXXXX (11 digits)
                dept = int(reg_no[4:7])  # NNN from position 4-6
                num = int(reg_no[7:])    # XXXX from position 7-10
                if num < 1 or num > 9999:
                    return False, "Student number must be between 0001 and 9999"
            elif length == 10:
                # Format: 99YYNNNXXX (10 digits)
                dept = int(reg_no[4:7])  # NNN from position 4-6
                num = int(reg_no[7:])    # XXX from position 7-9
                if num < 1 or num > 999:
                    return False, "Student number must be between 001 and 999"
            elif length == 9:
                # Format: 99YYNNNXX (9 digits)
                dept = int(reg_no[4:7])  # NNN from position 4-6
                num = int(reg_no[7:])    # XX from position 7-8
                if num < 1 or num > 99:
                    return False, "Student number must be between 01 and 99"
                    
        except (ValueError, IndexError):
            return False, "Invalid number format in registration number"
            
        return True, ""

    def _conn(self):
        return sqlite3.connect(self.db_path)

    def _init_db(self):
        with self._conn() as c:
            cur = c.cursor()
            
            # Create users table without is_first_login
            cur.execute("""
              CREATE TABLE IF NOT EXISTS users(
                  reg_no TEXT PRIMARY KEY,
                  password TEXT NOT NULL,
                  role TEXT NOT NULL
              )""")
              
            cur.execute("""
              CREATE TABLE IF NOT EXISTS students(
                  reg_no TEXT PRIMARY KEY,
                  name TEXT,
                  dept TEXT,
                  room_no TEXT,
                  father_name TEXT,
                  father_phone TEXT
              )""")
              
            # seed default accounts and student data
            default_accounts = [
                # Users table data
                ("INSERT OR IGNORE INTO users(reg_no,password,role) VALUES(?,?,?)", [
                    ("99220041142", "student@123", "student"),
                    ("99220040138", "student@123", "student"),
                    ("99220040139", "student@123", "student"),
                    ("PSBCSE", "faculty@123", "faculty")
                ]),
                # Students table data
                ("INSERT OR IGNORE INTO students(reg_no,name,dept,room_no,father_name,father_phone) VALUES(?,?,?,?,?,?)", [
                    ("99220041142", "Demo Student", "CSE", "A-101", "Mr. Kumar", "9876543210"),
                    ("99220040138", "John Doe", "CSE", "B-205", "Mr. Doe", "9876543211"),
                    ("99220040139", "Jane Smith", "CSE", "C-303", "Mr. Smith", "9876543212")
                ])
            ]
            
            # Execute all inserts
            for query, values in default_accounts:
                for value in values:
                    try:
                        cur.execute(query, value)
                    except sqlite3.Error as e:
                        print(f"Error inserting data: {e}")
            
            c.commit()

    def verify_user(self, reg_no, password):
        # First validate the registration number format
        valid, message = self.validate_reg_no(reg_no)
        if not valid:
            return None, message
        
        with self._conn() as c:
            cur = c.cursor()
            # First check if user exists
            cur.execute("SELECT role, password FROM users WHERE reg_no=?", (reg_no,))
            r = cur.fetchone()
            
            if not r:
                # For new valid student registration numbers, create account automatically
                if reg_no.startswith("99"):
                    try:
                        # Create new student account with default password
                        cur.execute("INSERT INTO users(reg_no, password, role) VALUES(?,?,?)",
                                  (reg_no, "student@123", "student"))
                        c.commit()
                        # If password matches default, let them in
                        if password == "student@123":
                            return "student", "Login successful"
                    except:
                        pass
                return None, "Invalid credentials. For new students, use registration number and password 'student@123'"
            
            stored_role, stored_password = r
            if password != stored_password:
                if password == "student@123":
                    return None, "Incorrect password. If you've changed your password, please use your new password."
                return None, "Incorrect password"
            
            return stored_role, "Login successful"

    def update_user_password(self, reg_no, old_password, new_password):
        with self._conn() as c:
            cur = c.cursor()
            cur.execute("SELECT password FROM users WHERE reg_no=?", (reg_no,))
            row = cur.fetchone()
            if not row:
                return False, "User not found"
                
            current_password = row[0]
            if current_password != old_password:
                return False, "Old password is incorrect"
                
            if new_password == "student@123":
                return False, "Cannot use the default password. Please choose a different password."
                
            if len(new_password) < 8:
                return False, "Password must be at least 8 characters long"
                
            # Update password
            cur.execute("""
                UPDATE users 
                SET password=?
                WHERE reg_no=?
            """, (new_password, reg_no))
            c.commit()
            return True, "Password updated successfully"

    def create_new_student(self, reg_no, name, dept, room_no="", father_name="", father_phone=""):
        """Create a new student with default password"""
        valid, message = self.validate_reg_no(reg_no)
        if not valid:
            return False, message
            
        with self._conn() as c:
            cur = c.cursor()
            try:
                # First create the user account with default password "student@123"
                cur.execute("INSERT INTO users(reg_no,password,role) VALUES(?,?,?)",
                          (reg_no, "student@123", "student"))
                          
                # Then create student details
                cur.execute("""
                    INSERT INTO students(reg_no,name,dept,room_no,father_name,father_phone) 
                    VALUES(?,?,?,?,?,?)
                """, (reg_no, name, dept, room_no, father_name, father_phone))
                c.commit()
                return True, f"Student {reg_no} created successfully. Default password is 'student@123'"
            except sqlite3.IntegrityError:
                return False, "Registration number already exists"
            except Exception as e:
                return False, f"Error creating student: {str(e)}"

    def update_student(self, reg_no, name, dept, room_no, father_name, father_phone):
        # Validate registration number
        valid, message = self.validate_reg_no(reg_no)
        if not valid:
            return False, message

        with self._conn() as c:
            cur = c.cursor()
            try:
                # Make sure user exists in users table with default password
                cur.execute("SELECT 1 FROM users WHERE reg_no=?", (reg_no,))
                if not cur.fetchone() and reg_no != "PSBCSE":
                    # Create user with default password if doesn't exist
                    cur.execute("INSERT INTO users(reg_no,password,role) VALUES(?,?,?)",
                              (reg_no, "student@123", "student"))
                
                # Update student details
                cur.execute("""
                    INSERT INTO students(reg_no,name,dept,room_no,father_name,father_phone) VALUES(?,?,?,?,?,?)
                    ON CONFLICT(reg_no) DO UPDATE SET
                    name=excluded.name, dept=excluded.dept, room_no=excluded.room_no,
                    father_name=excluded.father_name, father_phone=excluded.father_phone
                """, (reg_no, name, dept, room_no, father_name, father_phone))
                c.commit()
                return True, "Student details updated successfully"
            except Exception as e:
                return False, f"Error updating student: {str(e)}"

    def get_student(self, reg_no):
        """Get student details and create if doesn't exist"""
        with self._conn() as c:
            cur = c.cursor()
            # First check if the student exists
            cur.execute("""
                SELECT s.reg_no, s.name, s.dept, s.room_no, s.father_name, s.father_phone, u.role
                FROM students s
                LEFT JOIN users u ON s.reg_no = u.reg_no
                WHERE s.reg_no=?
            """, (reg_no,))
            r = cur.fetchone()
            
            if r:
                # Student exists, return their details
                return {
                    "reg_no": r[0],
                    "name": r[1] or '',
                    "dept": r[2] or '',
                    "room_no": r[3] or '',
                    "father_name": r[4] or '',
                    "father_phone": r[5] or '',
                    "role": r[6] or 'student'
                }
            
            # If student doesn't exist but reg_no is valid, create an empty record
            valid, _ = self.validate_reg_no(reg_no)
            if valid:
                try:
                    # Try to get any existing user record
                    cur.execute("SELECT role FROM users WHERE reg_no=?", (reg_no,))
                    user = cur.fetchone()
                    role = user[0] if user else 'student'
                    
                    # Create empty student record
                    cur.execute("""
                        INSERT OR IGNORE INTO students 
                        (reg_no, name, dept, room_no, father_name, father_phone)
                        VALUES (?, '', '', '', '', '')
                    """, (reg_no,))
                    
                    if not user:
                        # Create user account if doesn't exist
                        cur.execute("""
                            INSERT OR IGNORE INTO users (reg_no, password, role)
                            VALUES (?, 'student@123', 'student')
                        """, (reg_no,))
                    
                    c.commit()
                    return {
                        "reg_no": reg_no,
                        "name": '',
                        "dept": '',
                        "room_no": '',
                        "father_name": '',
                        "father_phone": '',
                        "role": role
                    }
                except sqlite3.Error:
                    pass
            return None

    def get_all_students(self):
        with self._conn() as c:
            cur = c.cursor()
            cur.execute("SELECT reg_no,name,dept,room_no,father_name,father_phone FROM students")
            rows = cur.fetchall()
            keys = ["reg_no","name","dept","room_no","father_name","father_phone"]
            return [dict(zip(keys, r)) for r in rows]

    def get_students_dict(self):
        return {s["reg_no"]: s for s in self.get_all_students()}

    def export_students_to_excel(self, filepath):
        """Export all students + login credentials to Excel file with formatting"""
        try:
            students = self.get_all_students()
            
            # Get login info for all students
            with self._conn() as c:
                cur = c.cursor()
                cur.execute("SELECT reg_no, password, role FROM users")
                login_rows = cur.fetchall()
                login_dict = {row[0]: {'password': row[1], 'role': row[2]} for row in login_rows}
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Students & Login"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add headers - including login info
            headers = ["Registration No", "Password", "Role", "Name", "Department", "Room No", "Father Name", "Father Phone"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Add student data with login info
            for row_num, student in enumerate(students, 2):
                reg_no = student.get('reg_no', '')
                login_info = login_dict.get(reg_no, {'password': 'student@123', 'role': 'student'})
                
                ws.cell(row=row_num, column=1).value = reg_no
                ws.cell(row=row_num, column=2).value = login_info.get('password', 'student@123')
                ws.cell(row=row_num, column=3).value = login_info.get('role', 'student')
                ws.cell(row=row_num, column=4).value = student.get('name', '')
                ws.cell(row=row_num, column=5).value = student.get('dept', '')
                ws.cell(row=row_num, column=6).value = student.get('room_no', '')
                ws.cell(row=row_num, column=7).value = student.get('father_name', '')
                ws.cell(row=row_num, column=8).value = student.get('father_phone', '')
                
                # Apply borders to data cells
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).border = border
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="left", vertical="center")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 15
            
            # Freeze first row
            ws.freeze_panes = "A2"
            
            # Save file
            wb.save(filepath)
            return True, f"Exported {len(students)} students with login credentials to Excel"
        except Exception as e:
            return False, f"Error exporting to Excel: {str(e)}"

    def import_students_from_excel(self, filepath):
        """Import/Update students from Excel file with login credentials - removes missing students and their logins"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Extract data from Excel (skip header row)
            excel_students = {}
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if row[0] is None:  # Skip empty rows
                    continue
                    
                reg_no = str(row[0]).strip()
                
                # Check if new format (with login columns) or old format
                if len(row) >= 8:  # New format with login info
                    password = str(row[1]).strip() if row[1] else 'student@123'
                    role = str(row[2]).strip() if row[2] else 'student'
                    name = str(row[3]).strip() if row[3] else ''
                    dept = str(row[4]).strip() if row[4] else ''
                    room_no = str(row[5]).strip() if row[5] else ''
                    father_name = str(row[6]).strip() if row[6] else ''
                    father_phone = str(row[7]).strip() if row[7] else ''
                else:  # Old format without login info
                    password = 'student@123'
                    role = 'student'
                    name = str(row[1]).strip() if row[1] else ''
                    dept = str(row[2]).strip() if row[2] else ''
                    room_no = str(row[3]).strip() if row[3] else ''
                    father_name = str(row[4]).strip() if row[4] else ''
                    father_phone = str(row[5]).strip() if row[5] else ''
                
                excel_students[reg_no] = {
                    'password': password,
                    'role': role,
                    'name': name,
                    'dept': dept,
                    'room_no': room_no,
                    'father_name': father_name,
                    'father_phone': father_phone
                }
            
            # Get current students and users from DB
            current_students = {s['reg_no']: s for s in self.get_all_students()}
            
            with self._conn() as c:
                cur = c.cursor()
                
                # Get current users
                cur.execute("SELECT reg_no FROM users")
                current_users = {row[0] for row in cur.fetchall()}
                
                # Add/Update students from Excel
                for reg_no, student_data in excel_students.items():
                    if reg_no in current_students:
                        # Update existing student
                        cur.execute("""
                            UPDATE students 
                            SET name=?, dept=?, room_no=?, father_name=?, father_phone=?
                            WHERE reg_no=?
                        """, (student_data['name'], student_data['dept'], student_data['room_no'],
                              student_data['father_name'], student_data['father_phone'], reg_no))
                    else:
                        # Add new student
                        cur.execute("""
                            INSERT INTO students(reg_no, name, dept, room_no, father_name, father_phone)
                            VALUES(?, ?, ?, ?, ?, ?)
                        """, (reg_no, student_data['name'], student_data['dept'], student_data['room_no'],
                              student_data['father_name'], student_data['father_phone']))
                    
                    # Add/Update login credentials
                    if reg_no in current_users:
                        # Update existing user
                        cur.execute("""
                            UPDATE users 
                            SET password=?, role=?
                            WHERE reg_no=?
                        """, (student_data['password'], student_data['role'], reg_no))
                    else:
                        # Create new user
                        cur.execute("""
                            INSERT INTO users(reg_no, password, role)
                            VALUES(?, ?, ?)
                        """, (reg_no, student_data['password'], student_data['role']))
                
                # Remove students and users not in Excel
                removed_count = 0
                removed_logins = 0
                for reg_no in current_students.keys():
                    if reg_no not in excel_students:
                        # Remove student record
                        cur.execute("DELETE FROM students WHERE reg_no=?", (reg_no,))
                        removed_count += 1
                        
                        # Remove login if exists
                        if reg_no in current_users:
                            cur.execute("DELETE FROM users WHERE reg_no=?", (reg_no,))
                            removed_logins += 1
                
                c.commit()
            
            return True, f"Imported {len(excel_students)} students with {len([s for s in excel_students.values() if s['role'] == 'student'])} active logins. Removed {removed_count} students and {removed_logins} logins."
        except Exception as e:
            return False, f"Error importing from Excel: {str(e)}"

    def delete_student(self, reg_no):
        """Delete a single student"""
        try:
            with self._conn() as c:
                cur = c.cursor()
                cur.execute("DELETE FROM students WHERE reg_no=?", (reg_no,))
                c.commit()
            return True, f"Student {reg_no} deleted successfully"
        except Exception as e:
            return False, f"Error deleting student: {str(e)}"

    def cleanup_old_students(self, photo_folder=None):
        """Delete students and their photos who registered more than 4.5 years ago"""
        current_date = datetime.now()
        current_year = current_date.year % 100
        current_month = current_date.month
        
        deleted_count = 0
        with self._conn() as c:
            cur = c.cursor()
            cur.execute("SELECT reg_no FROM students WHERE reg_no LIKE '99%'")
            rows = cur.fetchall()
            
            for row in rows:
                reg_no = row[0]
                try:
                    year = int(reg_no[2:4])
                    # Calculate difference in years
                    years_diff = current_year - year
                    # Adjust if current year wrapped around
                    if years_diff < 0:
                        years_diff += 100
                        
                    # Calculate approximate months (assuming joined in August)
                    months_diff = (years_diff * 12) + (current_month - 8)
                    
                    # 4.5 years = 54 months
                    if months_diff > 54:
                        # Delete from users
                        cur.execute("DELETE FROM users WHERE reg_no=?", (reg_no,))
                        # Delete from students
                        cur.execute("DELETE FROM students WHERE reg_no=?", (reg_no,))
                        deleted_count += 1
                        
                        # Delete photos if photo_folder is provided
                        if photo_folder and os.path.exists(photo_folder):
                            import glob
                            for photo_file in glob.glob(os.path.join(photo_folder, f"{reg_no}_*.jpg")):
                                try:
                                    os.remove(photo_file)
                                except Exception as e:
                                    print(f"Error deleting photo {photo_file}: {e}")
                except (ValueError, IndexError):
                    continue
            c.commit()
        return True, f"Cleaned up {deleted_count} old students (older than 4.5 years)"
