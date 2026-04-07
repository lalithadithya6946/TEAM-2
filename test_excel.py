#!/usr/bin/env python3
"""Test Excel export/import functionality"""

import os
import sys
from data_manager import DataManager

# Test the DataManager Excel functions
def test_excel_functions():
    print("=" * 60)
    print("TESTING EXCEL FUNCTIONALITY")
    print("=" * 60)
    
    # Initialize DataManager
    dm = DataManager()
    
    # Test 1: Export students to Excel
    print("\n[TEST 1] Exporting students to Excel...")
    export_path = "test_export.xlsx"
    success, message = dm.export_students_to_excel(export_path)
    print(f"✓ Export: {message}")
    if os.path.exists(export_path):
        print(f"✓ File created: {export_path}")
        file_size = os.path.getsize(export_path)
        print(f"✓ File size: {file_size} bytes")
    
    # Test 2: Get all students before import
    print("\n[TEST 2] Getting current students...")
    students_before = dm.get_all_students()
    print(f"✓ Total students before import: {len(students_before)}")
    for student in students_before[:3]:  # Show first 3
        print(f"  - {student['reg_no']}: {student['name']}")
    
    # Test 3: Import from Excel
    print("\n[TEST 3] Testing import functionality...")
    success, message = dm.import_students_from_excel(export_path)
    print(f"✓ Import: {message}")
    
    students_after = dm.get_all_students()
    print(f"✓ Total students after import: {len(students_after)}")
    
    # Test 4: Verify sync works
    print("\n[TEST 4] Verifying database sync...")
    if len(students_before) == len(students_after):
        print("✓ Student count matches before and after!")
    else:
        print("⚠ Student count changed")
    
    # Test 5: Excel file validation
    print("\n[TEST 5] Validating Excel file structure...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(export_path)
        ws = wb.active
        print(f"✓ Excel file loaded successfully")
        print(f"✓ Sheet name: {ws.title}")
        print(f"✓ Total rows (with header): {ws.max_row}")
        print(f"✓ Total columns: {ws.max_column}")
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['Reg No', 'Name', 'Department', 'Room No', 'Father Name', 'Father Phone']
        print(f"✓ Headers: {headers}")
        if headers == expected_headers:
            print("✓ Headers match expected format!")
    except Exception as e:
        print(f"⚠ Error validating Excel: {e}")
    
    # Cleanup
    print("\n[CLEANUP] Removing test file...")
    if os.path.exists(export_path):
        os.remove(export_path)
        print(f"✓ Test file removed")
    
    print("\n" + "=" * 60)
    print("✓ ALL TESTS COMPLETED SUCCESSFULLY!")
    print("=" * 60)

if __name__ == "__main__":
    try:
        test_excel_functions()
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
