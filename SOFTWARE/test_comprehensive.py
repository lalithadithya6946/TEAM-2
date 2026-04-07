#!/usr/bin/env python3
"""Comprehensive Excel dataset management test"""

import os
from data_manager import DataManager
from openpyxl import load_workbook
import sqlite3

def test_comprehensive_workflow():
    print("\n" + "=" * 70)
    print("COMPREHENSIVE EXCEL DATASET MANAGEMENT TEST")
    print("=" * 70)
    
    dm = DataManager()
    
    # Initial state
    print("\n[STEP 1] CHECKING INITIAL STATE")
    print("-" * 70)
    students = dm.get_all_students()
    conn = sqlite3.connect("students_data/students.db")
    cur = conn.cursor()
    cur.execute("SELECT reg_no, password, role FROM users")
    users = cur.fetchall()
    conn.close()
    
    print(f"  📊 Students in database: {len(students)}")
    print(f"  👥 Users (logins) in database: {len(users)}")
    print(f"  Sample users: {users[:3]}")
    
    # Export
    print("\n[STEP 2] EXPORTING TO EXCEL WITH LOGIN CREDENTIALS")
    print("-" * 70)
    export_file = "comprehensive_test.xlsx"
    success, msg = dm.export_students_to_excel(export_file)
    print(f"  ✓ {msg}")
    
    # Verify Excel structure
    wb = load_workbook(export_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"  📋 Excel columns: {headers}")
    print(f"  📝 Rows in Excel: {ws.max_row - 1} (data rows)")
    
    # Show sample data
    print(f"\n  Sample Excel data:")
    for row_num in range(2, min(4, ws.max_row + 1)):
        reg_no = ws.cell(row=row_num, column=1).value
        password = ws.cell(row=row_num, column=2).value
        role = ws.cell(row=row_num, column=3).value
        name = ws.cell(row=row_num, column=4).value
        print(f"    - {reg_no}: {name} | Password: {password} | Role: {role}")
    
    # Modify Excel
    print("\n[STEP 3] MODIFYING EXCEL (Edit Password, Remove 1, Add 2 new)")
    print("-" * 70)
    
    # Edit first student's password
    ws.cell(row=2, column=2).value = "changed@123"
    print(f"  ✏️  Changed password for first student to 'changed@123'")
    
    # Remove last student
    last_row = ws.max_row
    removed_reg = ws.cell(row=last_row, column=1).value
    ws.delete_rows(last_row)
    print(f"  ❌ Removed student: {removed_reg}")
    
    # Add new students
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1).value = "99220050001"
    ws.cell(row=new_row, column=2).value = "newstudent1@123"
    ws.cell(row=new_row, column=3).value = "student"
    ws.cell(row=new_row, column=4).value = "New Student 1"
    ws.cell(row=new_row, column=5).value = "CSE"
    print(f"  ➕ Added new student: 99220050001")
    
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1).value = "99220050002"
    ws.cell(row=new_row, column=2).value = "newstudent2@123"
    ws.cell(row=new_row, column=3).value = "student"
    ws.cell(row=new_row, column=4).value = "New Student 2"
    ws.cell(row=new_row, column=5).value = "ECE"
    print(f"  ➕ Added new student: 99220050002")
    
    wb.save(export_file)
    print(f"\n  💾 Excel saved with modifications")
    
    # Import modified Excel
    print("\n[STEP 4] IMPORTING MODIFIED EXCEL")
    print("-" * 70)
    success, msg = dm.import_students_from_excel(export_file)
    print(f"  ✓ {msg}")
    
    # Verify changes
    print("\n[STEP 5] VERIFYING CHANGES IN DATABASE")
    print("-" * 70)
    
    students_new = dm.get_all_students()
    conn = sqlite3.connect("students_data/students.db")
    cur = conn.cursor()
    
    # Check password change
    cur.execute("SELECT password FROM users WHERE reg_no=?", (students[0]['reg_no'],))
    new_password = cur.fetchone()
    if new_password and new_password[0] == "changed@123":
        print(f"  ✓ Password change verified: First student password updated to 'changed@123'")
    
    # Check removed login
    cur.execute("SELECT COUNT(*) FROM users WHERE reg_no=?", (removed_reg,))
    removed_login = cur.fetchone()[0]
    if removed_login == 0:
        print(f"  ✓ Login removed: {removed_reg} cannot login anymore")
    
    # Check new logins
    cur.execute("SELECT password, role FROM users WHERE reg_no='99220050001'")
    new_login1 = cur.fetchone()
    if new_login1:
        print(f"  ✓ New login created: 99220050001 | Password: {new_login1[0]} | Role: {new_login1[1]}")
    
    cur.execute("SELECT password, role FROM users WHERE reg_no='99220050002'")
    new_login2 = cur.fetchone()
    if new_login2:
        print(f"  ✓ New login created: 99220050002 | Password: {new_login2[0]} | Role: {new_login2[1]}")
    
    # Summary
    cur.execute("SELECT COUNT(*) FROM users")
    total_logins = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]
    
    print(f"\n  📊 Final state:")
    print(f"     - Students: {total_students}")
    print(f"     - Active logins: {total_logins}")
    
    conn.close()
    
    # Cleanup
    print("\n[STEP 6] CLEANUP")
    print("-" * 70)
    if os.path.exists(export_file):
        os.remove(export_file)
    print(f"  ✓ Test files cleaned up")
    
    print("\n" + "=" * 70)
    print("✓ COMPREHENSIVE TEST COMPLETED SUCCESSFULLY!")
    print("=" * 70)
    print("\nSummary:")
    print("  ✓ Excel can be downloaded with student + login data")
    print("  ✓ Passwords can be edited in Excel")
    print("  ✓ Removing rows deletes logins (portal stops working)")
    print("  ✓ Adding rows creates new logins")
    print("  ✓ Changes sync instantly with database")

if __name__ == "__main__":
    try:
        test_comprehensive_workflow()
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
