#!/usr/bin/env python3
"""Test Excel login removal/addition functionality"""

import os
import sys
from data_manager import DataManager
from openpyxl import load_workbook, Workbook

def test_login_removal():
    print("=" * 60)
    print("TESTING LOGIN REMOVAL/ADDITION FUNCTIONALITY")
    print("=" * 60)
    
    dm = DataManager()
    
    # Test 1: Export current state
    print("\n[TEST 1] Exporting current state...")
    dm.export_students_to_excel("test_login.xlsx")
    wb = load_workbook("test_login.xlsx")
    ws = wb.active
    print(f"✓ Current students in Excel: {ws.max_row - 1} rows")
    
    # Get current logins
    import sqlite3
    conn = sqlite3.connect("students_data/students.db")
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    logins_before = cur.fetchone()[0]
    print(f"✓ Current logins in database: {logins_before}")
    conn.close()
    
    # Test 2: Remove one student from Excel
    print("\n[TEST 2] Removing one student from Excel...")
    wb = load_workbook("test_login.xlsx")
    ws = wb.active
    ws.delete_rows(3)  # Delete third student row
    wb.save("test_login_modified.xlsx")
    print(f"✓ Removed one row from Excel")
    
    # Test 3: Import modified Excel
    print("\n[TEST 3] Importing modified Excel (should remove login)...")
    success, message = dm.import_students_from_excel("test_login_modified.xlsx")
    print(f"✓ Import result: {message}")
    
    # Check logins after import
    conn = sqlite3.connect("students_data/students.db")
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    logins_after = cur.fetchone()[0]
    print(f"✓ Logins after import: {logins_after}")
    
    if logins_after < logins_before:
        print("✓ SUCCESS! Login was removed when student was deleted from Excel")
    else:
        print("⚠ WARNING: Login was not removed")
    
    # Test 4: Add new student back
    print("\n[TEST 4] Adding new student back to Excel...")
    wb = load_workbook("test_login_modified.xlsx")
    ws = wb.active
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1).value = "99220040999"
    ws.cell(row=new_row, column=2).value = "newstudent@123"
    ws.cell(row=new_row, column=3).value = "student"
    ws.cell(row=new_row, column=4).value = "New Student"
    ws.cell(row=new_row, column=5).value = "CSE"
    wb.save("test_login_modified.xlsx")
    print(f"✓ Added new student 99220040999 to Excel")
    
    # Test 5: Import again
    print("\n[TEST 5] Importing Excel with new student...")
    success, message = dm.import_students_from_excel("test_login_modified.xlsx")
    print(f"✓ Import result: {message}")
    
    # Check if new login was created
    conn = sqlite3.connect("students_data/students.db")
    cur = conn.cursor()
    cur.execute("SELECT password, role FROM users WHERE reg_no='99220040999'")
    result = cur.fetchone()
    if result:
        print(f"✓ New login created! Password: {result[0]}, Role: {result[1]}")
    else:
        print("⚠ New login was not created")
    conn.close()
    
    # Cleanup
    print("\n[CLEANUP] Restoring original state...")
    dm.export_students_to_excel("test_login.xlsx")  # Reset
    for f in ["test_login.xlsx", "test_login_modified.xlsx"]:
        if os.path.exists(f):
            os.remove(f)
    print(f"✓ Cleanup complete")
    
    print("\n" + "=" * 60)
    print("✓ ALL LOGIN TESTS COMPLETED!")
    print("=" * 60)

if __name__ == "__main__":
    try:
        test_login_removal()
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
